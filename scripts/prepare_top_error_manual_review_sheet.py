"""Prepare a manual review sheet for AS-OCT ensemble top-error samples.

The script reads existing error-analysis outputs and writes review-ready CSV,
Excel, instructions, and a contact sheet. It does not modify predictions,
manifests, checkpoints, or training results.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from tempfile import TemporaryDirectory

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from PIL import Image, ImageOps


FRONT_COLUMNS = [
    "rank",
    "global_sample_id",
    "vault_label_um",
    "pred_ensemble_um",
    "signed_error_um",
    "abs_error_um",
    "auto_review_focus",
    "manual_label_check",
    "manual_image_quality",
    "manual_alignment_check",
    "manual_decision",
    "corrected_vault_um",
    "manual_comment",
]

MANUAL_COLUMNS = [
    "manual_label_check",
    "manual_image_quality",
    "manual_alignment_check",
    "manual_decision",
    "corrected_vault_um",
    "manual_comment",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Prepare manual review sheets for AS-OCT ensemble top-error samples."
    )
    parser.add_argument(
        "--top_error_csv",
        default="artifacts/reports/combined_batch_01_02/as_oct_error_analysis/as_oct_ensemble_top_error_samples.csv",
    )
    parser.add_argument(
        "--all_error_csv",
        default="artifacts/reports/combined_batch_01_02/as_oct_error_analysis/as_oct_ensemble_test_error_by_sample.csv",
    )
    parser.add_argument(
        "--review_index_csv",
        default="artifacts/reports/combined_batch_01_02/as_oct_error_analysis/top_error_review_package/top_error_review_index.csv",
    )
    parser.add_argument(
        "--image_dir",
        default="artifacts/reports/combined_batch_01_02/as_oct_error_analysis/top_error_review_package/images",
    )
    parser.add_argument(
        "--output_dir",
        default="artifacts/reports/combined_batch_01_02/as_oct_error_analysis/manual_review_sheet",
    )
    parser.add_argument("--top_k", type=int, default=10)
    return parser.parse_args()


def require_file(path: Path, description: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing {description}: {path}")


def normalize_path(value: object) -> Path | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().strip('"').strip("'")
    if not text or text.lower() in {"nan", "none"}:
        return None
    candidates = [Path(text), Path(text.replace("\\", "/"))]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def choose_image_path(row: pd.Series, image_dir: Path, output_dir: Path | None = None) -> Path | None:
    candidates: list[Path | None] = [
        normalize_path(row.get("copied_image_path", "")),
    ]
    rank = int(row.get("rank", 0) or 0)
    global_sample_id = str(row.get("global_sample_id", ""))
    search_dirs = [image_dir]
    if output_dir is not None:
        search_dirs.append(output_dir.parent / "top_error_review_package" / "images")

    for search_dir in search_dirs:
        if not search_dir.exists():
            continue
        if rank > 0:
            candidates.extend(sorted(search_dir.glob(f"rank{rank:02d}_*{safe_fragment(global_sample_id)}*")))
            candidates.extend(sorted(search_dir.glob(f"rank{rank:02d}_*")))
        candidates.extend(sorted(search_dir.glob(f"*{safe_fragment(global_sample_id)}*")))

    candidates.extend(
        [
            normalize_path(row.get("source_image_path", "")),
            normalize_path(row.get("oct_path", "")),
        ]
    )
    for candidate in candidates:
        if candidate is not None and candidate.exists():
            return candidate
    for candidate in candidates:
        if candidate is not None:
            return candidate
    return None


def read_image_for_preview(path: Path | str | None, max_size: tuple[int, int] = (900, 620)) -> tuple[Image.Image | None, dict[str, object]]:
    info: dict[str, object] = {
        "path": str(path) if path is not None else "",
        "exists": False,
        "mode": "",
        "size": "",
        "preview_success": False,
        "error": "",
    }
    resolved = normalize_path(path) if path is not None else None
    if resolved is None:
        info["error"] = "empty path"
        return None, info
    info["path"] = str(resolved)
    info["exists"] = resolved.exists()
    if not resolved.exists():
        info["error"] = "file not found"
        return None, info

    try:
        with Image.open(resolved) as src:
            src.load()
            info["mode"] = src.mode
            info["size"] = f"{src.width}x{src.height}"
            if src.mode == "RGB":
                img = src.copy()
            elif src.mode in {"L", "RGBA"}:
                img = src.convert("RGB")
            elif src.mode in {"I;16", "I;16B", "I;16L", "I", "F"}:
                arr = np.asarray(src).astype(np.float32)
                finite = np.isfinite(arr)
                if not finite.any():
                    raise ValueError("image array has no finite values")
                arr_min = float(np.nanmin(arr[finite]))
                arr_max = float(np.nanmax(arr[finite]))
                if arr_max <= arr_min:
                    arr8 = np.zeros(arr.shape, dtype=np.uint8)
                else:
                    arr8 = np.clip((arr - arr_min) / (arr_max - arr_min) * 255.0, 0, 255).astype(np.uint8)
                img = Image.fromarray(arr8, mode="L").convert("RGB")
            else:
                try:
                    img = src.convert("RGB")
                except Exception as exc:
                    raise ValueError(f"unsupported mode {src.mode}: {exc}") from exc

        img = ImageOps.autocontrast(img)
        img.thumbnail(max_size, Image.LANCZOS)
        info["preview_success"] = True
        return img, info
    except Exception as exc:
        info["error"] = f"PIL open failed: {exc}"
        return None, info


def compute_rank(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "rank" in df.columns:
        df["rank"] = pd.to_numeric(df["rank"], errors="coerce")
    elif "error_rank_desc" in df.columns:
        df["rank"] = pd.to_numeric(df["error_rank_desc"], errors="coerce")
    else:
        df = df.sort_values("abs_error_um", ascending=False).reset_index(drop=True)
        df["rank"] = np.arange(1, len(df) + 1)
    df["rank"] = df["rank"].fillna(pd.Series(np.arange(1, len(df) + 1), index=df.index)).astype(int)
    return df.sort_values("rank").reset_index(drop=True)


def find_copied_image(row: pd.Series, image_dir: Path) -> str:
    existing = str(row.get("copied_image_path", "") or "")
    normalized = normalize_path(existing)
    if normalized is not None and normalized.exists():
        return str(normalized)

    rank = int(row.get("rank", 0) or 0)
    global_sample_id = str(row.get("global_sample_id", ""))
    if rank > 0:
        matches = sorted(image_dir.glob(f"rank{rank:02d}_*{safe_fragment(global_sample_id)}*"))
        if matches:
            return str(matches[0])
        matches = sorted(image_dir.glob(f"rank{rank:02d}_*"))
        if matches:
            return str(matches[0])
    matches = sorted(image_dir.glob(f"*{safe_fragment(global_sample_id)}*"))
    if matches:
        return str(matches[0])
    return existing


def safe_fragment(value: str) -> str:
    return str(value).replace("\\", "_").replace("/", "_").replace(":", "_")


def build_auto_focus(row: pd.Series, repeated_patients: set[str]) -> str:
    tags: list[str] = []
    abs_error = float(row.get("abs_error_um", 0) or 0)
    signed_error = float(row.get("signed_error_um", 0) or 0)
    vault_range = str(row.get("vault_range", "")).lower()
    patient_key = str(row.get("global_patient_uid", "") or row.get("patient_uid", ""))
    label_qc = str(row.get("label_qc_flag", "")).strip().lower()

    if abs_error >= 250:
        tags.append("very_high_abs_error")
    elif abs_error >= 150:
        tags.append("high_abs_error")
    else:
        tags.append("moderate_abs_error")
    if vault_range == "low" and signed_error > 0:
        tags.append("low_vault_overestimation")
    if vault_range == "high" and signed_error < 0:
        tags.append("high_vault_underestimation")
    if patient_key and patient_key in repeated_patients:
        tags.append("repeated_patient_top_error")
    if label_qc and label_qc not in {"ok", "nan", "none"}:
        tags.append("label_qc_check_needed")
    return ";".join(tags)


def build_review_sheet(top_df: pd.DataFrame, review_index: pd.DataFrame | None, image_dir: Path, top_k: int) -> pd.DataFrame:
    df = compute_rank(top_df).head(top_k).copy()
    if review_index is not None and not review_index.empty and "global_sample_id" in review_index.columns:
        merge_cols = [col for col in ["global_sample_id", "copied_image_path", "source_image_path"] if col in review_index.columns]
        df = df.merge(review_index[merge_cols].drop_duplicates("global_sample_id"), on="global_sample_id", how="left", suffixes=("", "_review"))
        if "copied_image_path_review" in df.columns:
            df["copied_image_path"] = df.get("copied_image_path", "").fillna(df["copied_image_path_review"])
            df = df.drop(columns=["copied_image_path_review"])
        if "source_image_path" in df.columns and "oct_path" not in df.columns:
            df["oct_path"] = df["source_image_path"]

    patient_col = "global_patient_uid" if "global_patient_uid" in df.columns else "patient_uid"
    repeated = set(df[patient_col].dropna().astype(str).value_counts().loc[lambda x: x > 1].index) if patient_col in df.columns else set()

    df["copied_image_path"] = df.apply(lambda row: find_copied_image(row, image_dir), axis=1)
    df["image_exists"] = df["copied_image_path"].apply(lambda p: normalize_path(p) is not None and normalize_path(p).exists())

    widths = []
    heights = []
    for path in df["copied_image_path"]:
        resolved = normalize_path(path)
        if resolved is not None and resolved.exists():
            try:
                with Image.open(resolved) as img:
                    widths.append(img.width)
                    heights.append(img.height)
            except Exception:
                widths.append(np.nan)
                heights.append(np.nan)
        else:
            widths.append(np.nan)
            heights.append(np.nan)
    df["image_width"] = widths
    df["image_height"] = heights

    seed_cols = [col for col in ["pred_seed42_um", "pred_seed2026_um", "pred_seed3407_um"] if col in df.columns]
    if seed_cols:
        df["seed_pred_std_um"] = df[seed_cols].astype(float).std(axis=1, ddof=1)
        df["seed_pred_range_um"] = df[seed_cols].astype(float).max(axis=1) - df[seed_cols].astype(float).min(axis=1)
    else:
        df["seed_pred_std_um"] = np.nan
        df["seed_pred_range_um"] = np.nan

    df["auto_review_focus"] = df.apply(lambda row: build_auto_focus(row, repeated), axis=1)
    df["manual_label_check"] = "pending"
    df["manual_image_quality"] = "pending"
    df["manual_alignment_check"] = "pending"
    df["manual_decision"] = "pending"
    df["corrected_vault_um"] = ""
    df["manual_comment"] = ""

    wanted = [
        *FRONT_COLUMNS,
        "sample_id",
        "batch_id",
        "global_patient_uid",
        "patient_uid",
        "eye_side",
        "vault_range",
        "label_qc_flag",
        "measurement_ready_status",
        "pred_seed42_um",
        "pred_seed2026_um",
        "pred_seed3407_um",
        "seed_pred_std_um",
        "seed_pred_range_um",
        "oct_path",
        "copied_image_path",
        "image_exists",
        "image_width",
        "image_height",
    ]
    ordered = [col for col in wanted if col in df.columns]
    rest = [col for col in df.columns if col not in ordered]
    return df[ordered + rest]


def write_excel(sheet_df: pd.DataFrame, out_path: Path, thumbnail_dir: Path | None = None) -> tuple[bool, str]:
    try:
        # Some older openpyxl builds still reference deprecated NumPy aliases.
        if not hasattr(np, "float"):
            np.float = float  # type: ignore[attr-defined]
        if not hasattr(np, "int"):
            np.int = int  # type: ignore[attr-defined]
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        return False, f"openpyxl unavailable: {exc}"

    wb = Workbook()
    ws = wb.active
    ws.title = "manual_review"

    excel_df = sheet_df.copy()
    if "image_preview" not in excel_df.columns:
        insert_at = list(excel_df.columns).index("copied_image_path") + 1 if "copied_image_path" in excel_df.columns else len(excel_df.columns)
        excel_df.insert(insert_at, "image_preview", "")

    for col_idx, col_name in enumerate(excel_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, (_, row) in enumerate(excel_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(excel_df.columns, start=1):
            value = row[col_name]
            if pd.isna(value):
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)
        abs_error = float(row.get("abs_error_um", 0) or 0)
        if abs_error >= 250:
            fill = PatternFill("solid", fgColor="FCE4D6")
        elif abs_error >= 150:
            fill = PatternFill("solid", fgColor="FFF2CC")
        else:
            fill = None
        if fill:
            for col_idx in range(1, len(excel_df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_map = {
        "rank": 8,
        "global_sample_id": 34,
        "auto_review_focus": 42,
        "manual_label_check": 18,
        "manual_image_quality": 20,
        "manual_alignment_check": 24,
        "manual_decision": 18,
        "manual_comment": 36,
        "oct_path": 55,
        "copied_image_path": 55,
        "image_preview": 24,
    }
    for col_idx, col_name in enumerate(excel_df.columns, start=1):
        width = width_map.get(col_name, min(max(len(str(col_name)) + 2, 12), 22))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row_idx in range(2, len(excel_df) + 2):
        ws.row_dimensions[row_idx].height = 95

    embedded = 0
    failures: list[str] = []
    preview_col = list(excel_df.columns).index("image_preview") + 1
    for row_idx, (_, row) in enumerate(excel_df.iterrows(), start=2):
        path = normalize_path(row.get("chosen_image_path", "") or row.get("copied_image_path", ""))
        if path is None or not path.exists():
            failures.append(f"rank {row.get('rank', row_idx - 1)}: file not found")
            continue
        try:
            preview, info = read_image_for_preview(path, max_size=(260, 180))
            if preview is None:
                failures.append(f"rank {row.get('rank', row_idx - 1)}: {info.get('error')}")
                continue
            if thumbnail_dir is None:
                thumbnail_dir = out_path.parent / "_excel_preview_thumbnails"
            thumbnail_dir.mkdir(parents=True, exist_ok=True)
            thumb_path = thumbnail_dir / f"rank{int(row.get('rank', row_idx - 1)):02d}_preview.png"
            preview.save(thumb_path)
            img = XLImage(str(thumb_path))
            img.width = 130
            img.height = 90
            ws.add_image(img, f"{get_column_letter(preview_col)}{row_idx}")
            embedded += 1
        except Exception as exc:
            failures.append(f"rank {row.get('rank', row_idx - 1)}: {exc}")
            continue

    wb.save(out_path)
    if embedded == 0:
        return False, "Excel was created, but no image previews were embedded. " + "; ".join(failures[:5])
    message = f"Embedded {embedded} image previews."
    if failures:
        message += " Some previews failed: " + "; ".join(failures[:5])
    return embedded == len(excel_df), message


def write_instructions(out_path: Path) -> None:
    text = """# Top-error manual review instructions

## 目的

本 review sheet 用于逐个核对 AS-OCT seed ensemble 的 top-error samples，重点判断：

- POD1 vault label 是否录入正确；
- AS-OCT 输入图像是否为正确眼别、日期和 visit；
- 是否存在左右眼、日期或 visit 对齐问题；
- 是否存在明显图像质量问题；
- 是否需要将样本标记为 `label_suspected`、`image_quality_issue` 或 `alignment_issue`。

请优先检查 patient_052 双眼样本。如果双眼同时位于 top-error，需特别关注是否存在系统性的日期、visit、设备导出或 label 对齐问题。

## 字段填写建议

- `manual_label_check`: 填 `ok` / `suspected` / `corrected`。
- `manual_image_quality`: 填 `ok` / `poor` / `uncertain`。
- `manual_alignment_check`: 填 `ok` / `eye_side_suspected` / `date_suspected` / `visit_suspected`。
- `manual_decision`: 填 `keep` / `exclude` / `relabel` / `recheck`。
- `corrected_vault_um`: 仅当确认 label 需要修正时填写。
- `manual_comment`: 记录复查依据，例如图像质量、日期问题、左右眼疑似反转、POD1 label 疑似错误等。

## 判断标准

### label_suspected

如果 measurement crop 或人工记录显示 POD1 vault 与当前 `vault_label_um` 不一致，或同一眼多张 POD1 图差异异常且未解释，可标记为 label suspected。

### image_quality_issue

如果 AS-OCT 输入图像存在明显截断、模糊、扫描区域错误、伪影严重或无法判断关键解剖结构，可标记为 image quality issue。

### alignment_issue

如果输入图像的 eye side、exam date、visit 或 patient/sample 对应关系与 label 来源不一致，可标记为 alignment issue。优先核对左右眼和日期。

## 核对后保存

完成核对后，请另存为：

- `artifacts/reports/combined_batch_01_02/as_oct_error_analysis/manual_review_sheet/top_error_manual_review_sheet_checked.xlsx`

或：

- `artifacts/reports/combined_batch_01_02/as_oct_error_analysis/manual_review_sheet/top_error_manual_review_sheet_checked.csv`

不要覆盖原始 review sheet。
"""
    out_path.write_text(text, encoding="utf-8")


def make_contact_sheet(sheet_df: pd.DataFrame, out_path: Path) -> list[dict[str, object]]:
    n = len(sheet_df)
    cols = 2
    rows = int(np.ceil(n / cols))
    fig, axes = plt.subplots(rows, cols, figsize=(cols * 4.4, rows * 3.25))
    axes_arr = np.array(axes).reshape(-1)
    debug_rows: list[dict[str, object]] = []

    for ax, (_, row) in zip(axes_arr, sheet_df.iterrows()):
        path = normalize_path(row.get("chosen_image_path", "") or row.get("copied_image_path", ""))
        ax.axis("off")
        preview, info = read_image_for_preview(path, max_size=(1000, 720))
        info["rank"] = int(row.get("rank", 0) or 0)
        info["global_sample_id"] = row.get("global_sample_id", "")
        debug_rows.append(info)
        if preview is not None:
            ax.imshow(preview)
        else:
            ax.add_patch(plt.Rectangle((0, 0), 1, 1, transform=ax.transAxes, color="#e6e6e6"))
            reason = str(info.get("error", "unknown error"))
            if len(reason) > 58:
                reason = reason[:55] + "..."
            ax.text(
                0.5,
                0.55,
                "image unreadable",
                ha="center",
                va="center",
                transform=ax.transAxes,
                fontsize=10,
            )
            ax.text(
                0.5,
                0.42,
                reason,
                ha="center",
                va="center",
                transform=ax.transAxes,
                fontsize=7,
                color="#555555",
            )
        suffix = str(row["global_sample_id"]).replace("batch_01__", "b01__").replace("batch_02__", "b02__")
        title = (
            f"rank {int(row['rank']):02d} | {suffix}\n"
            f"label {float(row['vault_label_um']):.0f}, pred {float(row['pred_ensemble_um']):.0f}, "
            f"abs {float(row['abs_error_um']):.0f} um"
        )
        ax.set_title(title, fontsize=8)

    for ax in axes_arr[n:]:
        ax.axis("off")
    fig.tight_layout()
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return debug_rows


def main() -> None:
    args = parse_args()
    top_error_csv = Path(args.top_error_csv)
    review_index_csv = Path(args.review_index_csv)
    image_dir = Path(args.image_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    require_file(top_error_csv, "top-error CSV")
    if not Path(args.all_error_csv).exists():
        print(f"WARNING: all-error CSV not found; continuing without it: {args.all_error_csv}")
    if not review_index_csv.exists():
        print(f"WARNING: review index CSV not found; copied_image_path will be inferred when possible: {review_index_csv}")

    top_df = pd.read_csv(top_error_csv)
    review_index = pd.read_csv(review_index_csv) if review_index_csv.exists() else None
    sheet_df = build_review_sheet(top_df, review_index, image_dir, args.top_k)
    chosen_paths = []
    preflight_debug = []
    for _, row in sheet_df.iterrows():
        chosen = choose_image_path(row, image_dir, output_dir)
        chosen_paths.append(str(chosen) if chosen is not None else "")
        _, info = read_image_for_preview(chosen)
        info["rank"] = int(row.get("rank", 0) or 0)
        info["global_sample_id"] = row.get("global_sample_id", "")
        preflight_debug.append(info)
    sheet_df["chosen_image_path"] = chosen_paths
    sheet_df["image_exists"] = [bool(item.get("exists")) for item in preflight_debug]
    sheet_df["image_width"] = [
        int(str(item.get("size", "0x0")).split("x")[0]) if str(item.get("size", "")).count("x") == 1 and str(item.get("size", "0x0")).split("x")[0].isdigit() else np.nan
        for item in preflight_debug
    ]
    sheet_df["image_height"] = [
        int(str(item.get("size", "0x0")).split("x")[1]) if str(item.get("size", "")).count("x") == 1 and str(item.get("size", "0x0")).split("x")[1].isdigit() else np.nan
        for item in preflight_debug
    ]

    csv_path = output_dir / "top_error_manual_review_sheet.csv"
    xlsx_path = output_dir / "top_error_manual_review_sheet.xlsx"
    instructions_path = output_dir / "top_error_manual_review_instructions.md"
    contact_sheet_path = output_dir / "top_error_contact_sheet.png"

    sheet_df.to_csv(csv_path, index=False, encoding="utf-8")
    excel_embedded, excel_message = write_excel(sheet_df, xlsx_path, output_dir / "_excel_preview_thumbnails")
    write_instructions(instructions_path)
    contact_debug = make_contact_sheet(sheet_df, contact_sheet_path)

    found = int(sheet_df["image_exists"].sum()) if "image_exists" in sheet_df.columns else 0
    missing = int(len(sheet_df) - found)

    print(f"Top_k samples: {len(sheet_df)}")
    print("Image preview debug:")
    for item in contact_debug:
        print(
            "  "
            f"rank {int(item.get('rank', 0)):02d} | {item.get('global_sample_id')} | "
            f"path={item.get('path')} | exists={item.get('exists')} | "
            f"mode={item.get('mode')} | size={item.get('size')} | "
            f"preview_success={item.get('preview_success')} | error={item.get('error')}"
        )
    print(f"Read top-error samples: {len(sheet_df)}")
    print(f"Images found: {found}")
    print(f"Images missing: {missing}")
    print(f"Excel image embedding: {'success' if excel_embedded else 'not fully embedded'} - {excel_message}")
    print("Generated files:")
    for path in [csv_path, xlsx_path, instructions_path, contact_sheet_path]:
        print(f"  {path}")


if __name__ == "__main__":
    main()
